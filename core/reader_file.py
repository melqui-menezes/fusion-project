from openpyxl import load_workbook
from unidecode import unidecode

def csv_to_array(file):
        wb = load_workbook(file)
        ws = wb.active
        search_head = []
        head = [
        		'CENTRAL', 'COOPERATIVA', 'AGENCIA', 'DATA_CONTRATACAO', 'CPF_VENDEDOR', 'SITUACAO',
        		'CELULAR', 'EMAIL', 'COMISSAO', 'NUM_PROPOSTA', 'NUM_APOLICE', 'TIPO_ASSINATURA'
        		]
        data = {}
        client = {}
        data_list = []

        for row in ws.iter_rows(max_row=1, values_only=True):
                for value in row:
                        search_head.append(unidecode(value.replace(" ","_")))

        for row in ws.iter_rows(min_row=2, values_only=True):
                x = 0
                for value in row:
                        if search_head[x] in head:
                                client = {search_head[x]:value}
                                data.update(client)
                                client = dict(data)
                        x += 1
                data_list.append(client)
        return data_list
'''
global_file = 'Global.xlsx'
global_list = csv_to_array(global_file)

individual_file = 'Individual.xlsx'
individual_list = csv_to_array(individual_file)
'''
